import requests
import sys
import numpy as np
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
from openpyxl import load_workbook


load_wb = load_workbook("mystock.xlsx", data_only=True)

load_ws = load_wb['Sheet1']

dates = ['20200401','20200402','20200403','20200404','20200405','20200406','20200407','20200408','20200409','20200410','20200411','20200412','20200413','20200414','20200415','20200416','20200417','20200418','20200419','20200420','20200421','20200422','20200423','20200424','20200425']
# data = ""
j = 2
name = {"name": 0}
for date in dates:
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
    data = requests.get('http://www.genie.co.kr/chart/top200?ditc=D&ymd='+date+'&hh=23&rtm=N&pg=1', headers=headers)

    soup = BeautifulSoup(data.text, 'html.parser')
    songs = soup.select('#body-content > div.newest-list > div > table > tbody > tr')
    load_ws.cell(1, j, date)

    i = 2
    for song in songs:

        rank = song.select_one('td.number').text[0:20].strip()
        title = song.select_one('td.info > a.title.ellipsis').text.strip()
        artist = song.select_one('td.info > a.title.ellipsis').text
        if title in name:
            name[title] += 1
        else:
            name[title] = 1
        load_ws.cell(i, j, title+'-'+artist)
        i += 1

    j += 1

load_wb.save("mystock.xlsx")
song_title = []
N = len(name)
# value = name.values()
song_value = name.values()
song_title = list(zip(name.keys())) # dict_keys to list
ind = np.arange(N)

print(song_value)
plt.bar(ind, song_value)

plt.ylabel('횟수')
plt.title('Geine month song')
# plt.xticks(song_title, rotation=45)
plt.show()
