import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

load_wb = load_workbook("mystock.xlsx", data_only=True)
#시트이름 불러오기
load_ws = load_wb['Sheet1']

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&ymd=20200425&hh=23&rtm=N&pg=1',headers=headers)



#html을 검색하기 용이하게 하기 위해 BeautifulSoup를 사용
soup = BeautifulSoup(data.text, 'html.parser')

# html에서 tr을 select로 불러오기
songs = soup.select('#body-content > div.newest-list > div > table > tbody > tr')

#시작하는 행을 지정
i = 2

#movies들의 반복문
for song in songs:
    #td중 클래스가 number인 녀석의 텍스트 중 0,1번째 글자를 가져오고 좌우 공백을 제거
    rank = song.select_one('td.number').text[0:2].strip()
    #info 아래 title, ellipsis가 붙은 a태그를 불러옴
    title = song.select_one('td.info > a.title.ellipsis').text.strip()
    #info아래 artist, ellipsis가 붙은 a태그와 a태그의 텍스트를 가져옴
    artist = song.select_one('td.info > a.artist.ellipsis').text
    
    #cell에 값 입력
    load_ws.cell(i, 1, rank)
    load_ws.cell(i, 2, title)
    load_ws.cell(i, 3, artist)
    
    i += 1

    print(rank, title, artist)

#엑셀 파일에 저장
load_wb.save("mystock.xlsx")


