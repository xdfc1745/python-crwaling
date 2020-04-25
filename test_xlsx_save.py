from openpyxl import load_workbook

load_wb = load_workbook("mymusic.xlsx", data_only = True)

load_ws = load_wb['Sheet1']

print(load_ws.cell(1, 1).value)

load_ws.cell(5, 7, '5행7열')
load_wb.save("mymusic.xlsx")